import io
import operator
import random
import string

import openpyxl
import pandas as pd
from io import BytesIO

import datetime

from PIL import Image
from openpyxl.styles import PatternFill

from admin.models import PickerSettingsModel, PickerHistoryModel
from database import s3, DefaultRepository
from orders.models import OrdersAccountModel, OrdersAddressModel, OrdersOrderModel
from orgs.models import OrganizationModel
from payments.models import BalanceHistoryModel
from products.models import ProductModel, ProductSizeModel
from products.repository import ProductsRepository
from products.utils import parse_wildberries_card
from strings import *

file_types = {
    'image': {
        'types': ['jpeg', 'png', 'jpg', 'webp'],
        'max': 1048576 * 10,
    },
    'video': {
        'types': ['mp4', 'mov'],
        'max': 1048576 * 50,
    },
    'other': {
        'types': ['pdf'],
        'max': 104857 * 10,
    }
}


def generate_filename():
    alphanumeric_characters = string.ascii_letters + string.digits
    return ''.join(random.choice(alphanumeric_characters) for _ in range(32))


async def verify_file(file, available_media: list):
    if file.size == 0:
        raise Exception(string_storage_empty_file)

    fileinfo = file.filename.rsplit('.', maxsplit=1)
    if len(fileinfo) != 2:
        raise Exception(string_storage_wrong_filetype)

    filename = fileinfo[0]
    filetype = fileinfo[1]

    if len(filename) == 0:
        raise Exception(string_storage_empty_filename)

    available_types = sum([file_types[t]['types'] for t in available_media], [])

    if filetype.lower() not in available_types:
        raise Exception(string_storage_wrong_filetype + f'. Только: {available_types}')

    for t in file_types:
        if filetype in file_types[t]['types']:
            if file.size > file_types[t]['max']:
                raise Exception(string_storage_max_size)


async def s3_save(file_bytes, file_name, file_type):
    if file_type in file_types['image']['types']:
        image = Image.open(io.BytesIO(file_bytes))
        image.save(io.BytesIO(), format='WEBP', optimize=True, quality=15)

        webp_bytes = io.BytesIO()
        image.save(webp_bytes, format='WebP')
        file_content = webp_bytes.getvalue()
        file_type = 'webp'

    else:
        file_content = file_bytes

    s3.upload_fileobj(io.BytesIO(file_content), 'greedybear', f'{file_name}.{file_type}')

    return file_name, file_type


def set_type(value, field_type):
    if field_type == 'INTEGER':
        return int(value) if value else None

    elif field_type == 'VARCHAR':
        return str(value) if value else None

    elif field_type == 'BOOLEAN':
        return bool(value)

    elif field_type == 'DATETIME':
        return datetime.datetime.fromisoformat(value) if value else None

    elif field_type == 'DATE':
        return datetime.date.fromisoformat(value) if value else None

    elif field_type == 'TIME':
        return datetime.datetime.strptime(value, '%H:%M:%S').time() if value else None

    elif field_type == 'FLOAT':
        return float(value) if value else None

    return value


def detect_date(string_date):
    result = None
    if str(string_date) != 'nan' and string_date:

        if len(string_date) == 19:
            date_format = '%Y-%m-%d %H:%M:%S'
        else:
            string_date = string_date.replace(' ', '')
            date_format = '%d.%m.%Y'

        result = datetime.datetime.strptime(string_date, date_format)

    return result


async def force_save_product(wb_article, wb_title, wb_size_origName, wb_size_optionId, org_id):
    try:
        wb_url = f'https://www.wildberries.ru/catalog/{wb_article}/detail.aspx'

        title, p_article, sizes, picture_data = parse_wildberries_card(wb_url)

        filename = generate_filename() if picture_data else None

        await ProductsRepository.create_product(
            {
                'org_id': int(org_id),
                'wb_article': wb_article,
                'wb_title': wb_title,
                'status': 1,
                'media': filename + '.webp',
            },
            [size.model_dump() for size in sizes]
        )

        if picture_data:
            await s3_save(picture_data, filename, 'webp')

        return True

    except Exception as e:

        await ProductsRepository.create_product(
            {
                'org_id': int(org_id),
                'wb_article': wb_article,
                'wb_title': wb_title,
                'status': 1,
            },
            [{
                'wb_size_origName': wb_size_origName,
                'wb_size_optionId': wb_size_optionId,
                'wb_in_stock': False,
                'wb_price': None,
                'barcode': None,
                'is_active': True,
            }]
        )

        return False


async def force_save_order(x):
    db_sizes = await DefaultRepository.get_records(ProductSizeModel)
    db_products = await DefaultRepository.get_records(ProductModel)
    db_organizations = await DefaultRepository.get_records(OrganizationModel)

    df_products = pd.DataFrame([{'id': x.id, 'org_id': x.org_id, 'article': x.wb_article} for x in db_products])
    df_organizations = pd.DataFrame([{'id': x.id, 'title': x.title} for x in db_organizations])
    df_sizes = pd.DataFrame([{'id': x.id, 'product_id': x.product_id, 'wb_size_origName': x.wb_size_origName,
                              'wb_size_name': x.wb_size_name} for x in db_sizes])
    for xx in x:
        if force_save:

            query = df_organizations.query(f"title == '{line['organization_title']}'")
            organization_id = query['id'].iloc[0] if len(query.values) != 0 else None

            if not organization_id:
                logs_orders.append(
                    {
                        'target': line['order_uuid'],
                        'success': False,
                        'detail': 'Организация не найдена',
                        'value': line['organization_title'],
                        'line': line['line_number'],
                        'orders_type': orders_type,
                        'server': server.name,
                    }
                )
                continue

            query = df_products.query(f"article == '{line['product_article']}'")
            product_id = query['id'].iloc[0] if len(query.values) != 0 else None

            if not product_id:
                logs_orders.append(
                    {
                        'target': line['order_uuid'],
                        'success': False,
                        'detail': f'Товар не найден в списке товаров {line["organization_title"]}',
                        'value': line['product_article'],
                        'line': line['line_number'],
                        'orders_type': orders_type,
                        'server': server.name,
                    }
                )
                continue

            query = df_sizes.query(f'product_id == {product_id}')

            if len(query) == 0:
                logs_orders.append(
                    {
                        'target': line['order_uuid'],
                        'success': False,
                        'detail': f'Размеры товара не найдены',
                        'value': line['product_article'],
                        'line': line['line_number'],
                        'orders_type': orders_type,
                        'server': server.name,
                    }
                )
                continue

            size_id = query['id'].iloc[0]

            if len(query) != 1:
                for i, row in query.iterrows():
                    if (row['wb_size_origName'] == line['product_size'] or
                            row['wb_size_name'] == line['product_size']):
                        size_id = row['id']
                        break

            data_orders_to_db.append(
                {
                    'wb_keyword': line['product_title'],
                    'wb_price': int(line['price']),
                    'wb_uuid': line['order_uuid'],
                    'wb_status': line['status'],
                    'wb_collect_code': line['collect_code'],

                    'status': 3,
                    'description': 'forced',

                    'dt_planed': detect_date(line['dt_ordered']),
                    'dt_ordered': detect_date(line['dt_ordered']),
                    'dt_delivered': detect_date(line['dt_delivered']),
                    'dt_collected': detect_date(line['dt_collected']),

                    'size_id': int(size_id),
                    'account_id': int(account_id)
                }
            )

            logs_orders.append(
                {
                    'target': line['order_uuid'],
                    'success': True,
                    'detail': 'Заказ сохранен методом ForceSave',
                    'value': line['order_uuid'],
                    'line': line['line_number'],
                    'orders_type': orders_type,
                    'server': server.name,
                }
            )
            continue


async def parse_excel_lines(excel_lines, columns):
    return [
        {
            **{v: str(excel_line[n]) if str(excel_line[n]) != 'nan' else None for v, n in columns.items()},
            'line_number': line_number + 2,
        }
        for line_number, excel_line in enumerate(excel_lines)
    ]


async def refresh_active_and_collected(data, servers):
    db_accounts = await DefaultRepository.get_records(OrdersAccountModel)
    db_addresses = await DefaultRepository.get_records(OrdersAddressModel)
    db_orders = await DefaultRepository.get_records(OrdersOrderModel)

    df_accounts = pd.DataFrame([{'id': x.id, 'number': x.number} for x in db_accounts])
    df_addresses = pd.DataFrame([{'id': x.id, 'address': x.address} for x in db_addresses])
    df_orders = pd.DataFrame([{'id': x.id, 'size_id': x.size_id, 'wb_uuid': x.wb_uuid, 'dt_collected': x.dt_collected}
                              for x in db_orders])

    excel_columns = {
        'address': 0,
        'status': 1,
        'account_name': 2,
        'telnum': 3,
        'collect_code': 4,
        'organization_title': 5,
        'product_article': 6,
        'product_size': 7,
        'product_title': 8,
        'price': 9,
        'dt_ordered': 10,
        'dt_delivered': 11,
        'dt_collected': 12,
        'account_number': 13,
        'order_uuid': 14,
    }

    logs_accounts = []
    logs_orders = []

    processed_accounts = []
    data_orders_to_db = []

    for server in servers:

        data_accounts_to_db = []

        active_orders_content = await data[f'active-{server.id}'].read()
        collected_orders_content = await data[f'collected-{server.id}'].read()

        orders = {
            'active': await parse_excel_lines(
                pd.read_excel(BytesIO(active_orders_content), dtype=str).values.tolist(), excel_columns),
            'collected': await parse_excel_lines(
                pd.read_excel(BytesIO(collected_orders_content), dtype=str).values.tolist(), excel_columns),
        }

        # Process accounts
        for orders_type, lines in orders.items():

            for line in reversed(lines):

                print(server.name, line['line_number'], 'a')

                if line['account_number'] in processed_accounts:
                    continue

                if not line['account_number']:
                    logs_accounts.append(
                        {
                            'target': line['account_number'],
                            'success': False,
                            'detail': 'Номер аккаунта не указан',
                            'value': line['account_number'],
                            'line': line['line_number'],
                            'orders_type': orders_type,
                            'server': server.name,
                        }
                    )
                    continue

                if not line['address']:
                    logs_accounts.append(
                        {
                            'target': line['account_number'],
                            'success': False,
                            'detail': 'Адрес ПВЗ не указан',
                            'value': line['address'],
                            'line': line['line_number'],
                            'orders_type': orders_type,
                            'server': server.name,
                        }
                    )
                    continue

                if not line['account_name']:
                    logs_accounts.append(
                        {
                            'target': line['account_number'],
                            'success': False,
                            'detail': 'Имя аккаунта не указано',
                            'value': line['account_name'],
                            'line': line['line_number'],
                            'orders_type': orders_type,
                            'server': server.name,
                        }
                    )
                    continue

                query = df_accounts.query(f"number == '{line['account_number']}'")
                account_id = query['id'].iloc[0] if len(query.values) != 0 else None

                if not account_id:
                    query = df_addresses.query(f"address == '{line['address']}'")
                    address_id = query['id'].iloc[0] if len(query.values) != 0 else None

                    if not address_id:
                        logs_accounts.append(
                            {
                                'target': line['account_number'],
                                'success': False,
                                'detail': 'Адрес ПВЗ не найден в базе данных',
                                'value': line['address'],
                                'line': line['line_number'],
                                'orders_type': orders_type,
                                'server': server.name,
                            }
                        )
                        continue

                    data_accounts_to_db.append(
                        {
                            'number': line['account_number'],
                            'name': line['account_name'],
                            'is_active': True,
                            'address_id': address_id,
                            'server_id': server.id,
                        }
                    )

                    logs_accounts.append(
                        {
                            'target': line['account_number'],
                            'success': True,
                            'detail': 'Аккаунт добавлен в базу данных',
                            'value': line['account_number'],
                            'line': line['line_number'],
                            'orders_type': orders_type,
                            'server': server.name,
                        }
                    )

                processed_accounts.append(line['account_number'])

        # Save accounts
        if data_accounts_to_db:
            await DefaultRepository.save_records([{'model': OrdersAccountModel, 'records': data_accounts_to_db}])
            db_accounts = await DefaultRepository.get_records(OrdersAccountModel)
            df_accounts = pd.DataFrame([{'id': x.id, 'number': x.number} for x in db_accounts])

        # Process orders
        for orders_type, lines in orders.items():

            for line in reversed(lines):

                print(server.name, line['line_number'], 'o')

                if not line['order_uuid']:
                    logs_orders.append(
                        {
                            'target': line['order_uuid'],
                            'success': False,
                            'detail': 'UUID заказа не указан',
                            'value': line['order_uuid'],
                            'line': line['line_number'],
                            'orders_type': orders_type,
                            'server': server.name,
                        }
                    )
                    continue

                query = df_orders.query(f"wb_uuid == '{line['order_uuid']}'")
                order_id = query['id'].iloc[0] if len(query.values) != 0 else None

                if order_id:

                    if orders_type == 'collected' and query['dt_collected'].iloc[0]:
                        continue

                    query = df_accounts.query(f"number == '{line['account_number']}'")
                    account_id = query['id'].iloc[0] if len(query.values) != 0 else None

                    if not account_id:
                        logs_orders.append(
                            {
                                'target': line['order_uuid'],
                                'success': False,
                                'detail': 'Аккаунт заказа не найден',
                                'value': line['account_number'],
                                'line': line['line_number'],
                                'orders_type': orders_type,
                                'server': server.name,
                            }
                        )
                        continue

                    data_orders_to_db.append(
                        {
                            'id': order_id,
                            'wb_status': line['status'],
                            'wb_collect_code': line['collect_code'],
                            'dt_ordered': detect_date(line['dt_ordered']),
                            'dt_delivered': detect_date(line['dt_delivered']),
                            'dt_collected': detect_date(line['dt_collected']),
                            'account_id': account_id,
                        }
                    )

                    logs_orders.append(
                        {
                            'target': line['order_uuid'],
                            'success': True,
                            'detail': 'Данные заказа обновлены в базе данных',
                            'value': line['order_uuid'],
                            'line': line['line_number'],
                            'orders_type': orders_type,
                            'server': server.name,
                        }
                    )
                    continue

                logs_orders.append(
                    {
                        'target': line['order_uuid'],
                        'success': False,
                        'detail': 'Заказ не найден по UUID',
                        'value': line['order_uuid'],
                        'line': line['line_number'],
                        'orders_type': orders_type,
                        'server': server.name,
                    }
                )

    # Save orders
    if data_orders_to_db:
        await DefaultRepository.save_records([{'model': OrdersOrderModel, 'records': data_orders_to_db}])

    return {'accounts': logs_accounts, 'orders': logs_orders}


async def identify_orders_xlsx(data):
    with_payments = True if data['with_payments'] == 'true' else False
    is_test = True if data['is_test'] == 'true' else False

    excel_columns = {
        'sid': 15,
        'status': 10,
        'uuid': 12,
        'price': 14,
    }

    xlsx_orders_content = await data['orders'].read()
    xlsx_orders = await parse_excel_lines(pd.read_excel(
        BytesIO(xlsx_orders_content), dtype=str).values.tolist(), excel_columns)

    db_orders = await DefaultRepository.get_records(
        OrdersOrderModel,
        select_related=[OrdersOrderModel.size, OrdersOrderModel.account],
        deep_related=[
            [OrdersOrderModel.size, ProductSizeModel.product],
        ],
        joins=[ProductSizeModel, ProductModel],
    )

    df_orders = pd.DataFrame([
        {
            'id': x.id,
            'wb_price': x.wb_price,
            'org_id': x.size.product.org_id,
        }
        for x in db_orders
    ])

    logs = []

    records_orders = []
    records_payments = []

    for line in xlsx_orders:

        if not line['sid']:
            logs.append(
                {
                    'target': line['sid'],
                    'success': False,
                    'detail': 'Не указан SID заказа',
                    'value': line['sid'],
                    'line': line['line_number'],
                }
            )
            continue

        if not line['status']:
            logs.append(
                {
                    'target': line['sid'],
                    'success': False,
                    'detail': 'Не указан статус заказа',
                    'value': line['status'],
                    'line': line['line_number'],
                }
            )
            continue

        if not line['uuid']:
            logs.append(
                {
                    'target': line['sid'],
                    'success': False,
                    'detail': 'Не указан UUID заказа',
                    'value': line['uuid'],
                    'line': line['line_number'],
                }
            )
            continue

        if not line['price']:
            logs.append(
                {
                    'target': line['sid'],
                    'success': False,
                    'detail': 'Не указана стоимость заказа',
                    'value': line['price'],
                    'line': line['line_number'],
                }
            )
            continue

        try:
            int(line['price'])
        except:
            logs.append(
                {
                    'target': line['sid'],
                    'success': False,
                    'detail': 'Цена должна быть числом',
                    'value': line['price'],
                    'line': line['line_number'],
                }
            )
            continue

        query = df_orders.query(f"id == {line['sid']}")

        order_id = query['id'].iloc[0] if len(query.values) != 0 else None
        wb_price = query['wb_price'].iloc[0] if len(query.values) != 0 else None
        org_id = query['org_id'].iloc[0] if len(query.values) != 0 else None

        if not order_id:
            logs.append(
                {
                    'target': line['sid'],
                    'success': False,
                    'detail': 'Заказ не найден по SID',
                    'value': line['sid'],
                    'line': line['line_number'],
                }
            )
            continue

        if not wb_price:
            logs.append(
                {
                    'target': line['sid'],
                    'success': False,
                    'detail': 'Не указана стоимость заказа в базе данных',
                    'value': wb_price,
                    'line': line['line_number'],
                }
            )
            continue

        if not org_id:
            logs.append(
                {
                    'target': line['sid'],
                    'success': False,
                    'detail': 'Организация заказа не опознана',
                    'value': org_id,
                    'line': line['line_number'],
                }
            )
            continue

        if 'Все артикулы заказаны' in line['status']:
            status = 3
            wb_status = 'Все артикулы заказаны'
            description = line['status']

            try:
                dt_ordered = datetime.datetime.strptime(
                    line['status'].strip().split('Все артикулы заказаны')[1].strip(),
                    '%d.%m.%Y %H:%M:%S'
                )
            except:
                logs.append(
                    {
                        'target': line['sid'],
                        'success': False,
                        'detail': 'Ошибка обработки даты заказа. Требуемый формат: ДД.ММ.ГГГГ ЧЧ:ММ:СС',
                        'value': line['status'],
                        'line': line['line_number'],
                    }
                )
                continue

        else:
            status = 4
            wb_status = 'Отмена'
            description = line['status']
            dt_ordered = None

        uuid = line['uuid'].replace(' ', '')
        if ',' in line['uuid']:
            uuid = line['uuid'].replace(' ', '').split(',')[-1]

        # Обновление заказа
        updated_order = {
            'id': order_id,
            'status': status,
            'wb_status': wb_status,
            'wb_uuid': uuid,
            'description': description,
            'dt_ordered': dt_ordered,
        }

        if with_payments:
            updated_order['wb_price'] = int(line['price'])
        records_orders.append(updated_order)

        # Возврат заморозки
        records_payments.append(
            {
                'description': f'Разморозка стоимости товара по задаче №{order_id}',
                'amount': int(wb_price),
                'org_id': org_id,
                'action_id': 4
            }
        )

        # Списание заказа
        records_payments.append(
            {
                'description': f'Оплата стоимости товара по задаче №{order_id}',
                'amount': int(line['price']),
                'org_id': org_id,
                'action_id': 3
            }
        )

        logs.append(
            {
                'target': line['uuid'],
                'success': True,
                'detail': 'Заказн обновлен, оплата совершена' if (with_payments or is_test) else 'Заказ обновлен',
                'value': f'Возврат: {wb_price}; Списание: {line["price"]}' if (with_payments or is_test) else None,
                'line': line['line_number'],
            }
        )

    for item in logs:
        if not item['success']:
            return logs

    records_to_db = [{'model': OrdersOrderModel, 'records': records_orders}]

    if with_payments:
        records_to_db.append({'model': BalanceHistoryModel, 'records': records_payments})

    if not is_test:
        await DefaultRepository.save_records(records_to_db)

    return logs


async def generate_plan_xlsx(servers, bad_accounts):
    class ScheduleCell:
        dt: datetime.datetime
        org: str | None

        def __init__(self, dt) -> None:
            self.dt = dt
            self.org = None

    def calc_cells(dts, dte, osd):
        return int(
            (dte - dts).total_seconds() /
            osd.total_seconds()
        )

    async def proc_0(org_accs, used_addrs):
        new_org_accs = []

        for i in range(0, len(org_accs)):

            # K - amount of orders planned on account's address for today
            K = used_addrs.count(org_accs[i]['address_id'])
            org_accs[i]['K'] = K

            # L = K + W
            org_accs[i]['L'] = K + org_accs[i]['W']

            # Check L
            if org_accs[i]['L'] in range(org_accs[i]['contractor'].load_l_min,
                                         org_accs[i]['contractor'].load_l_max + 1):
                new_org_accs.append(org_accs[i])

        return new_org_accs

    async def proc_1(accs, settings):
        maxH = max(acc['H'] for acc in accs)
        maxH = 100 if maxH == 0 else maxH

        for acc in accs:

            acc['X'] = ((maxH - acc['H']) / maxH) * 100
            acc['Z'] = 0
            acc['AB'] = acc['L']
            acc['AD'] = settings.r4 - acc['T']

            if acc['M'] is not None:

                if acc['H'] == 0:
                    acc['Z'] = settings.r2

                else:

                    t = (datetime.datetime.now() - acc['M']).days

                    if t > settings.r2:
                        acc['Z'] = settings.r2 + t * settings.r3

                    else:
                        acc['Z'] = t

            else:
                acc['Z'] = settings.r2 + 10

            if acc['L'] == 0:
                acc['AB'] = 0

            elif acc['L'] >= settings.r4:
                acc['AB'] = settings.r4 - 1

            if acc['T'] == 0:
                acc['AD'] = 0

        return accs

    async def proc_2(accs, settings):
        maxX = max(acc['X'] for acc in accs)
        maxZ = max(acc['Z'] for acc in accs)

        for acc in accs:
            if not acc['M']:
                maxZ = settings.r2 + 10
                break

        maxX = 100 if maxX == 0 else maxX
        maxZ = 100 if maxZ == 0 else maxZ

        for acc in accs:
            acc['Y'] = (acc['X'] / maxX) * 100
            acc['AA'] = (acc['Z'] / maxZ) * 100
            acc['AC'] = (acc['AB'] / (settings.r4 - 1)) * 100
            acc['AE'] = acc['AD'] / settings.r4 * 100

        return accs

    async def proc_3(accs, settings):
        maxY = max(acc['Y'] for acc in accs)
        maxAA = max(acc['AA'] for acc in accs)
        maxAC = max(acc['AC'] for acc in accs)
        maxAE = max(acc['AE'] for acc in accs)

        for acc in accs:
            acc['AF'] = (
                    (
                            acc['Y'] * settings.l2 +
                            acc['AA'] * settings.l3 +
                            acc['AC'] * settings.l4 +
                            acc['AE'] * settings.l5
                    ) * 100 /
                    (
                            maxY * settings.l2 +
                            maxAA * settings.l3 +
                            maxAC * settings.l4 +
                            maxAE * settings.l5
                    )
            )

        return accs

    bad_accounts = (
        bad_accounts.replace('\n', '').replace(' ', '').split('\r')
        if bad_accounts
        else []
    )

    db_settings = await DefaultRepository.get_records(PickerSettingsModel)
    db_settings = db_settings[0]

    db_orders = await DefaultRepository.get_records(
        OrdersOrderModel,
        filters=[OrdersOrderModel.dt_ordered.isnot(None)],
        select_related=[OrdersOrderModel.size, OrdersOrderModel.account],
        deep_related=[
            [OrdersOrderModel.size, ProductSizeModel.product],
            [OrdersOrderModel.size, ProductSizeModel.product, ProductModel.organization]
        ],
        joins=[ProductSizeModel, ProductModel, OrganizationModel],
    )

    for order in db_orders:
        if not order.account:
            raise Exception(f'Не указан аккаунт у заказа #{order.id}')

    df_orders = pd.DataFrame(
        {
            'account_id': order.account_id,
            'address_id': order.account.address_id,
            'org_id': order.size.product.org_id,
            'org': order.size.product.organization.title,
            'dt_ordered': order.dt_ordered,
            'dt_delivered': order.dt_delivered,
            'dt_collected': order.dt_collected,
        }
        for order in db_orders
    )

    settings_last_order = datetime.timedelta(days=db_settings.lo)
    settings_account_life = datetime.timedelta(days=db_settings.al)
    settings_now = datetime.datetime.now()
    settings_now_date = datetime.date.today()
    settings_k_format = db_settings.k_format

    used_addresses = []
    total_selected_accounts = []

    for server in servers:

        db_tasks = await DefaultRepository.get_records(
            model=OrdersOrderModel,
            filters=[OrdersOrderModel.status == 2],
            filtration=[OrganizationModel.server_id == server.id],
            select_related=[OrdersOrderModel.size],
            deep_related=[
                [OrdersOrderModel.size, ProductSizeModel.product],
                [OrdersOrderModel.size, ProductSizeModel.product, ProductModel.organization]
            ],
            joins=[ProductSizeModel, ProductModel, OrganizationModel],
        )

        selected_accs, used_addrs, accs = [], used_addresses.copy(), []
        used_accs = {c.contractor.name: [] for c in server.contractors}

        logs_book = openpyxl.Workbook()
        result_book = openpyxl.Workbook()

        ws_res = result_book.active
        ws_res['B1'] = 'артикул'
        ws_res['C1'] = 'кол-во'
        ws_res['D1'] = 'размер'
        ws_res['E1'] = 'ключевой запрос'
        ws_res['F1'] = 'вход'
        ws_res['G1'] = 'ПВЗ'
        ws_res['H1'] = 'подсказка'
        ws_res['I1'] = 'порядок'
        ws_res['K1'] = 'статус'
        ws_res['L1'] = 'аккаунт'
        ws_res['M1'] = 'id заказа'
        ws_res['N1'] = 'ип'
        ws_res['O1'] = 'цена'
        ws_res['P1'] = 'sid'
        ws_res['Z2'] = settings_now.strftime('%d.%m')

        line = 1
        for order in db_tasks:
            line += 1

            size_name = ''
            if order.size.wb_size_name:
                size_name = order.size.wb_size_name
            if order.size.wb_size_origName:
                size_name = order.size.wb_size_origName

            ws_res[f'A{line}'] = line - 1
            ws_res[f'B{line}'] = order.size.product.wb_article
            ws_res[f'C{line}'] = 1
            ws_res[f'D{line}'] = size_name
            ws_res[f'E{line}'] = order.wb_keyword
            ws_res[f'N{line}'] = order.size.product.organization.title
            ws_res[f'O{line}'] = order.size.wb_price // 100
            ws_res[f'P{line}'] = order.id

        names = [(order.size.product.organization.title, order.size.product.wb_article) for order in db_tasks]

        arts = {}
        orgs = {}

        for org, art in names:
            if orgs.get(org):
                orgs[org] += 1

            else:
                orgs[org] = 1
                arts[org] = {}

            if arts[org].get(art):
                arts[org][art] += 1

            else:
                arts[org][art] = 1

        orgs = sorted(orgs.items(), key=operator.itemgetter(1), reverse=True)
        orgs = dict(orgs)

        ################################################################################################################
        # SCHEDULE
        ################################################################################################################

        logs = logs_book.create_sheet('Расписание')
        logs['A1'] = 'Количество задач'
        logs['A2'] = 'Количество ячеек'
        logs['A3'] = 'Шаг ячеек (мин)'
        logs['A4'] = 'Время начала'
        logs['A5'] = 'Время окончания'
        logs['A6'] = 'Размер созданного расписния'
        logs['A7'] = 'Время первой ячейки'
        logs['A8'] = 'Время последней ячейки'
        logs['A9'] = 'Название расписания'
        logs['B9'] = server.schedule.title

        tasks_amount = sum(orgs.values())
        max_one_step_duration = server.schedule.time_max_min_per_step
        min_one_step_duration = server.schedule.time_min_min_per_step

        start_time = datetime.datetime.combine(settings_now_date, server.schedule.time_start)
        end_time = datetime.datetime.combine(settings_now_date, server.schedule.time_end)

        middle_point = datetime.datetime.combine(settings_now_date, server.schedule.time_first_point)
        last_point = datetime.datetime.combine(settings_now_date, server.schedule.time_second_point)

        # end time point
        one_step_duration = datetime.timedelta(minutes=max_one_step_duration)
        cells_amount = calc_cells(start_time, end_time, one_step_duration)

        # middle point
        if cells_amount < tasks_amount:
            end_time = middle_point
            cells_amount = calc_cells(start_time, end_time, one_step_duration)

            # middle point + dynamic
            if cells_amount < tasks_amount:

                one_step_duration = datetime.timedelta(minutes=(cells_amount / tasks_amount) * max_one_step_duration)
                one_step_duration = one_step_duration - datetime.timedelta(microseconds=one_step_duration.microseconds)

                # last point
                if one_step_duration < datetime.timedelta(minutes=min_one_step_duration):

                    one_step_duration = datetime.timedelta(minutes=max_one_step_duration)
                    end_time = last_point
                    cells_amount = calc_cells(start_time, end_time, one_step_duration)

                    # last point + dynamic
                    if cells_amount < tasks_amount:
                        one_step_duration = datetime.timedelta(
                            minutes=(cells_amount / tasks_amount) * max_one_step_duration)
                        one_step_duration = one_step_duration - datetime.timedelta(
                            microseconds=one_step_duration.microseconds)
                        cells_amount = calc_cells(start_time, end_time, one_step_duration)

                        if one_step_duration < datetime.timedelta(minutes=min_one_step_duration):
                            logs['A10'] = f'Недостаточно ячеек расписания для {server.name}'

                            logs_bytes = io.BytesIO()

                            logs_book.save(logs_bytes)

                            logs_bytes.seek(0)

                            logs_filename, logs_filetype = await s3_save(logs_bytes.getvalue(), generate_filename(),
                                                                         'xlsx')

                            await DefaultRepository.save_records([
                                {
                                    'model': PickerHistoryModel, 'records':
                                    [
                                        {
                                            'logs': f'{logs_filename}.{logs_filetype}',
                                            'server_id': server.id,
                                        }
                                    ]
                                }
                            ])

                            continue

        logs['B1'] = tasks_amount
        logs['B2'] = cells_amount
        logs['B3'] = one_step_duration
        logs['B4'] = start_time
        logs['B5'] = end_time

        # Creating schedule
        schedule = []
        current_time = start_time
        while current_time < end_time:
            schedule.append(ScheduleCell(current_time))
            current_time += one_step_duration

        if schedule[-1].dt > end_time - one_step_duration:
            schedule = schedule[:-1]

        logs['B6'] = f'{len(schedule)} ячеек'
        logs['B7'] = f'{schedule[0].dt}'
        logs['B8'] = f'{schedule[-1].dt}'

        time_length = (end_time - start_time).total_seconds()

        # Assigning organizations to schedule
        for org in orgs:

            amount = orgs[org]
            single_length = time_length / amount
            assigned_amount = 0

            for x in range(amount):
                dt = start_time + datetime.timedelta(seconds=x * single_length)

                # Find closest to dt free cell in schedule
                closest_cell = None

                for cell in schedule:
                    if cell.dt >= dt and not cell.org:
                        closest_cell = cell
                        break

                if closest_cell:
                    assigned_amount += 1
                    closest_cell.org = org

        used_dt = []
        used_lines = []

        for org in arts:
            m = max(arts[org].values())

            for i in range(m):

                for art in arts[org]:

                    if arts[org][art] != 0:

                        line = 1
                        for row in ws_res.rows:

                            if str(row[1].value) == str(art) and str(row[13].value) == str(
                                    org) and line not in used_lines:

                                for cell in schedule:
                                    if cell.org and cell.org == org and cell.dt not in used_dt:
                                        used_dt.append(cell.dt)
                                        row[10].value = cell.dt.strftime(settings_k_format)
                                        break

                                used_lines.append(line)
                                break

                            line += 1

                        arts[org][art] -= 1

        ################################################################################################################
        # COMMON POOL
        ################################################################################################################
        db_accounts = await DefaultRepository.get_records(
            OrdersAccountModel,
            filters=[OrdersAccountModel.server_id == server.id],
            select_related=[OrdersAccountModel.address],
        )

        common_pool = []
        logs = logs_book.create_sheet('Общий пул')
        line = 1
        logs[f'A{line}'] = 'ID аккаунта'
        logs[f'B{line}'] = 'Номер'
        logs[f'C{line}'] = 'Имя'
        logs[f'D{line}'] = 'Статус аккаунта'
        logs[f'E{line}'] = 'ID адреса'
        logs[f'F{line}'] = 'Адрес'
        logs[f'G{line}'] = 'Район'
        logs[f'H{line}'] = 'Курьер'
        logs[f'I{line}'] = 'Статус адреса'
        logs[f'J{line}'] = 'Исключен'
        logs[f'K{line}'] = 'T'
        logs[f'L{line}'] = 'W'

        for account in db_accounts:

            line += 1
            logs[f'A{line}'] = account.id
            logs[f'B{line}'] = account.number
            logs[f'C{line}'] = account.name
            logs[f'D{line}'] = account.number in bad_accounts or account.is_active
            logs[f'E{line}'] = account.address.id
            logs[f'F{line}'] = account.address.address
            logs[f'G{line}'] = account.address.district
            logs[f'H{line}'] = account.address.contractor
            logs[f'I{line}'] = account.address.is_active

            # account active, account address active, account number not in bad_accounts
            if account.number in bad_accounts:
                logs[f'J{line}'] = 'Исключен, реестровый аккаунт'
                continue

            if not account.is_active:
                logs[f'J{line}'] = 'Исключен, аккаунт не активен'
                continue

            if not account.address.is_active:
                logs[f'J{line}'] = 'Исключен, адрес аккаунта не активен'
                continue

            # account last order
            query = (
                df_orders
                .query(f"account_id == {account.id} and dt_collected.isnull()")
                .sort_values('dt_ordered', ascending=True)
                .head(1)
            )

            last_order = query['dt_ordered'].iloc[0] if len(query.values) != 0 else None

            if last_order and last_order + settings_last_order < settings_now_date:
                logs[f'J{line}'] = \
                    f'Исключен, последний заказ на аккаунте позднее {settings_last_order.days} дней с текущей даты'
                continue

            # account registration date
            query = (
                df_orders
                .query(f"account_id == {account.id}")
                .sort_values('dt_ordered', ascending=True)
                .head(1)
            )

            default_reg_date = query['dt_ordered'].iloc[0] if len(query.values) != 0 else settings_now_date
            reg_date = account.reg_date if account.reg_date else default_reg_date

            if reg_date + settings_account_life < settings_now_date:
                logs[
                    f'J{line}'] = f'Исключен, с даты регистрации аккаунта {reg_date} прошло ' \
                                  f'{(settings_now_date - reg_date).days} ' \
                                  f'дней с текущей даты (макс допустимо {settings_account_life.days} дней)'
                continue

            # Contractor
            contractor = next(
                (
                    server_contractor for server_contractor in server.contractors
                    if server_contractor.contractor.id == account.address.contractor_id
                ), None
            )

            if not contractor:
                logs[f'J{line}'] = 'Исключен, курьер адреса аккаунта отсутствует на сервере'
                continue

            # T
            T = (df_orders.query(f'account_id == {account.id} and dt_collected.isnull()').shape[0])

            if T not in range(contractor.load_t_min, contractor.load_t_max + 1):
                logs[f'J{line}'] = 'Исключен, выход за допустимый интервал кол-ва активных заказов на аккаунте'
                continue

            # W
            W = (df_orders.query(f'address_id == {account.address_id} and dt_collected.isnull()').shape[0])

            common_pool.append({
                'account_id': account.id,
                'address_id': account.address_id,
                'number': account.number,
                'address': account.address.address,
                'contractor': contractor,
                'district': account.address.district,

                'T': T,  # amount of active orders on account
                'W': W,  # amount of active orders on account's address
            })

            logs[f'J{line}'] = 'Не исключен'
            logs[f'K{line}'] = T
            logs[f'L{line}'] = W

        ################################################################################################################
        # ACCOUNTS PICKER
        ################################################################################################################

        for org in orgs:
            amount = orgs[org]

            org_used_accs, org_used_addr, org_accs = [], [], []

            cs_tmp = sorted(server.contractors, key=lambda x: x.load_percent, reverse=True)
            contractors = []
            remaining = amount

            for c_tmp in cs_tmp[:-1]:

                cmax = round(amount * c_tmp.load_percent)

                if cmax >= remaining:
                    cmax = remaining
                    remaining = 0

                elif cmax < remaining:
                    remaining -= cmax

                contractors.append({
                    'name': c_tmp.contractor.name,
                    'usages': [],
                    'max': cmax
                })

            contractors.append({
                'name': cs_tmp[-1].contractor.name,
                'usages': [],
                'max': remaining
            })

            ws_org = logs_book.create_sheet(org)
            line = 1
            ws_org[f'A{line}'] = 'ID аккаунта'
            ws_org[f'B{line}'] = 'Номер'
            ws_org[f'C{line}'] = 'ID адреса'
            ws_org[f'D{line}'] = 'Адрес'
            ws_org[f'E{line}'] = 'Курьер'
            ws_org[f'F{line}'] = 'Район'
            ws_org[f'G{line}'] = 'Организация'
            ws_org[f'H{line}'] = 'Исключен'
            ws_org[f'I{line}'] = 'T'
            ws_org[f'J{line}'] = 'W'
            ws_org[f'K{line}'] = 'I'
            ws_org[f'L{line}'] = 'M'
            ws_org[f'M{line}'] = 'H'
            ws_org[f'N{line}'] = 'K'
            ws_org[f'O{line}'] = 'L'
            ws_org[f'P{line}'] = 'X'
            ws_org[f'Q{line}'] = 'Z'
            ws_org[f'R{line}'] = 'AB'
            ws_org[f'S{line}'] = 'AD'
            ws_org[f'T{line}'] = 'Y'
            ws_org[f'U{line}'] = 'AA'
            ws_org[f'V{line}'] = 'AC'
            ws_org[f'W{line}'] = 'AE'
            ws_org[f'X{line}'] = 'AF (Итоговый)'
            ws_org[f'Y{line}'] = 'Выбран'

            for acc in common_pool:

                line += 1

                ws_org[f'A{line}'] = acc['account_id']
                ws_org[f'B{line}'] = acc['number']
                ws_org[f'C{line}'] = acc['address_id']
                ws_org[f'D{line}'] = acc['address']
                ws_org[f'E{line}'] = acc['contractor'].contractor.name
                ws_org[f'F{line}'] = acc['district']
                ws_org[f'G{line}'] = org
                ws_org[f'I{line}'] = acc['T']
                ws_org[f'J{line}'] = acc['W']

                I = (df_orders.query(f"account_id == {acc['account_id']} and org == '{org}'").shape[0])
                ws_org[f'K{line}'] = I

                if I != 0:
                    ws_org[f'H{line}'] = 'Исключен, ИП уже заказывал на этот аккаунт'
                    continue

                query = (
                    df_orders
                    .query(f"address_id == {acc['address_id']} and org == '{org}'")
                    .sort_values('dt_ordered', ascending=False)
                    .head(1)
                )

                M = query['dt_ordered'].iloc[0] if len(query.values) != 0 else None

                if M:
                    ws_org[f'L{line}'] = M.strftime('%d.%m.%Y')
                else:
                    ws_org[f'L{line}'] = 'Не найдено'

                if M and M > acc['contractor'].load_m:
                    ws_org[f'H{line}'] = 'Исключен, дата последнего заказа вне допустимого интервала'
                    continue

                H = (df_orders.query(f"address_id == {acc['address_id']} and org == '{org}'").shape[0])

                ws_org[f'M{line}'] = H
                org_accs.append({
                    'org': org,
                    **acc,
                    'I': I,
                    'H': H,  # amount of all orders of org on address of account
                    'M': M,  # date of the last purchase of org on account's address
                    'logs': line,
                })

                ws_org[f'H{line}'] = 'Не исключен'

            for i in range(0, amount):

                org_accs = await proc_0(org_accs, used_addrs)  # L | K W

                if not org_accs:
                    break

                org_accs = await proc_1(org_accs, db_settings)  # X Z AB AD   | H L T
                org_accs = await proc_2(org_accs, db_settings)  # Y AA AC AE  | X Z AB AD
                org_accs = await proc_3(org_accs, db_settings)  # AF          | Y AA AC AE

                org_accs = sorted(org_accs, key=lambda x: x['AF'], reverse=True)

                if i == 0:
                    for oa in org_accs:
                        ws_org[f'N{oa["logs"]}'] = oa['K']
                        ws_org[f'O{oa["logs"]}'] = oa['L']
                        ws_org[f'P{oa["logs"]}'] = oa['X']
                        ws_org[f'Q{oa["logs"]}'] = oa['Z']
                        ws_org[f'R{oa["logs"]}'] = oa['AB']
                        ws_org[f'S{oa["logs"]}'] = oa['AD']
                        ws_org[f'T{oa["logs"]}'] = oa['Y']
                        ws_org[f'U{oa["logs"]}'] = oa['AA']
                        ws_org[f'V{oa["logs"]}'] = oa['AC']
                        ws_org[f'W{oa["logs"]}'] = oa['AE']
                        ws_org[f'X{oa["logs"]}'] = oa['AF']

                for org_acc in org_accs:

                    if org_acc['address_id'] in org_used_addr:
                        continue

                    # J
                    for contractor in contractors:
                        if (len(contractor['usages']) != contractor['max'] and
                                org_acc['contractor'].contractor.name == contractor['name']):

                            HH = (
                                df_orders.query(f'address_id == {org_acc["address_id"]} and dt_collected.isnull()').loc[
                                :,
                                'account_id'].unique().shape[0])

                            aa_accs = []
                            JJ = 0
                            for aacc in total_selected_accounts:
                                if aacc['address_id'] == org_acc['address_id'] and aacc['account_id'] not in aa_accs:
                                    aa_accs.append(aacc['account_id'])
                                    JJ += 1

                            II = HH + JJ  # кол во аккаунтов с активными заказами на адресе аккаунта

                            if II >= org_acc['contractor'].load_i and org_acc['T'] == 0:
                                continue

                            J = used_accs[org_acc['contractor'].contractor.name].count(org_acc['account_id'])

                            if J in range(org_acc['contractor'].load_j_min, org_acc['contractor'].load_j_max + 1):
                                org_acc['J'] = J
                                contractor['usages'].append(org_acc['account_id'])

                                used_addrs.append(org_acc['address_id'])
                                used_accs[org_acc['contractor'].contractor.name].append(org_acc['account_id'])

                                org_used_addr.append(org_acc['address_id'])
                                selected_accs.append(org_acc)

                                total_selected_accounts.append(org_acc)

                                ws_org[f'Y{org_acc["logs"]}'] = 'Выбран'
                                fill = PatternFill(patternType='solid', fgColor='FF00FF00')
                                ws_org[f'D{org_acc["logs"]}'].fill = fill

                                break

            s = 0
            ws_org[f'AA1'] = f'Всего'
            ws_org[f'AC1'] = f'из {amount}'

            for i in range(0, len(contractors)):
                t = len(contractors[i]['usages'])
                ws_org[f'AA{i + 2}'] = contractors[i]['name']
                ws_org[f'AB{i + 2}'] = t
                ws_org[f'AC{i + 2}'] = contractors[i]['max']
                s += t

            ws_org[f'AB1'] = f'{s}'

        random.shuffle(selected_accs)
        line = 1
        for row in ws_res.rows:
            line += 1
            for acc in selected_accs:

                if acc['org'] == ws_res[f'N{line}'].value and not acc.get('selected'):
                    acc['selected'] = True
                    ws_res[f'G{line}'] = acc['address']
                    ws_res[f'L{line}'] = acc['number']
                    break

        ws_res[f'Z2'] = datetime.datetime.now().strftime('%d.%m')

        result_bytes = io.BytesIO()
        logs_bytes = io.BytesIO()

        result_book.save(result_bytes)
        logs_book.save(logs_bytes)

        result_bytes.seek(0)
        logs_bytes.seek(0)

        result_filename, result_filetype = await s3_save(result_bytes.getvalue(), generate_filename(), 'xlsx')
        logs_filename, logs_filetype = await s3_save(logs_bytes.getvalue(), generate_filename(), 'xlsx')

        await DefaultRepository.save_records([
            {
                'model': PickerHistoryModel, 'records':
                [
                    {
                        'result': f'{result_filename}.{result_filetype}',
                        'logs': f'{logs_filename}.{logs_filetype}',
                        'server_id': server.id,
                    }
                ]
            }
        ])

        used_addresses = used_addrs.copy()
