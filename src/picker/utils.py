import datetime
import io
from io import BytesIO

import openpyxl
import pandas as pd

from database import Repository
from gutils import Strings
from orders.models import OrdersOrderModel, OrdersAccountModel, OrdersAddressModel
from orgs.models import OrganizationModel
from payments.models import BalanceHistoryModel
from picker.models import PickerOrderStatus, PickerHistoryModel
from products.models import ProductSizeModel, ProductModel


## MACROSES

def detect_date(string_date):
    result = None
    if str(string_date) != 'nan' and string_date:

        if len(string_date) == 19:
            date_format = '%Y-%m-%d %H:%M:%S'
        else:
            string_date = string_date.replace(' ', '')
            date_format = '%d.%m.%Y'

        result = datetime.datetime.strptime(string_date, date_format)

    return result


async def parse_excel_lines(excel_lines, columns):
    return [
        {
            **{v: str(excel_line[n]) if str(excel_line[n]) != 'nan' else None for v, n in columns.items()},
            'line_number': line_number + 2,
        }
        for line_number, excel_line in enumerate(excel_lines)
    ]


async def refresh_active_and_collected(data, servers):
    # Caching data from db
    db_accounts = await Repository.get_records(OrdersAccountModel)
    db_addresses = await Repository.get_records(OrdersAddressModel)
    db_statuses = await Repository.get_records(PickerOrderStatus)
    db_orders = await Repository.get_records(
        OrdersOrderModel,
        select_related=[OrdersOrderModel.size, OrdersOrderModel.account, OrdersOrderModel.picker_status],
        deep_related=[
            [OrdersOrderModel.size, ProductSizeModel.product],
        ],
        joins=[ProductSizeModel, ProductModel],
    )

    df_accounts = pd.DataFrame([{'id': x.id, 'number': x.number} for x in db_accounts])
    df_addresses = pd.DataFrame([{'id': x.id, 'address': x.address} for x in db_addresses])
    df_orders = pd.DataFrame([
        {
            'id': x.id,
            'wb_price': x.wb_price,
            'sys_status': x.status,
            'size_id': x.size_id,
            'org_id': x.size.product.org_id,
            'wb_uuid': x.wb_uuid,
            'dt_collected': x.dt_collected,
            'is_success': x.picker_status.is_success if x.picker_status else None
        }
        for x in db_orders
    ])

    ac_excel_columns = {
        'address': 0,
        'status': 1,
        'account_name': 2,
        'telnum': 3,
        'collect_code': 4,
        'organization_title': 5,
        'product_article': 6,
        'product_size': 7,
        'product_title': 8,
        'price': 9,
        'dt_ordered': 10,
        'dt_delivered': 11,
        'dt_collected': 12,
        'account_number': 13,
        'uuid': 14,
    }

    p_excel_columns = {
        'sid': 15,
        'status': 10,
        'uuid': 12,
    }

    logs_accounts = []
    logs_orders = []
    logs_payments = []

    processed_accounts = []
    processed_orders = []

    data_orders_to_db = []
    data_payments_to_db = []

    for server in servers:

        data_accounts_to_db = []

        # Reading xlsx files
        active_orders_content = await data[f'active-{server.id}'].read()
        collected_orders_content = await data[f'collected-{server.id}'].read()
        plan_orders_content = await data[f'plan-{server.id}'].read()

        orders = {
            'active': await parse_excel_lines(
                pd.read_excel(BytesIO(active_orders_content), dtype=str).values.tolist(), ac_excel_columns),
            'collected': await parse_excel_lines(
                pd.read_excel(BytesIO(collected_orders_content), dtype=str).values.tolist(), ac_excel_columns),
        }

        plan_orders = await parse_excel_lines(
            pd.read_excel(BytesIO(plan_orders_content), dtype=str).values.tolist(), p_excel_columns)

        for plan_line in plan_orders:
            if plan_line['uuid'] and ',' in plan_line['uuid']:
                plan_line['uuid'] = plan_line['uuid'].split(',')[-1].strip()

        # Process accounts
        for orders_type, lines in orders.items():

            for line in reversed(lines):

                if line['account_number'] in processed_accounts:
                    continue

                if not line['account_number']:
                    logs_accounts.append(
                        {
                            'target': line['account_number'],
                            'success': False,
                            'detail': 'Номер аккаунта не указан',
                            'value': line['account_number'],
                            'line': line['line_number'],
                            'orders_type': orders_type,
                            'server': server.name,
                        }
                    )
                    continue

                if not line['address']:
                    logs_accounts.append(
                        {
                            'target': line['account_number'],
                            'success': False,
                            'detail': 'Адрес ПВЗ не указан',
                            'value': line['address'],
                            'line': line['line_number'],
                            'orders_type': orders_type,
                            'server': server.name,
                        }
                    )
                    continue

                if not line['account_name']:
                    logs_accounts.append(
                        {
                            'target': line['account_number'],
                            'success': False,
                            'detail': 'Имя аккаунта не указано',
                            'value': line['account_name'],
                            'line': line['line_number'],
                            'orders_type': orders_type,
                            'server': server.name,
                        }
                    )
                    continue

                query = df_accounts.query(f"number == '{line['account_number']}'")
                account_id = query['id'].iloc[0] if len(query.values) != 0 else None

                if not account_id:
                    query = df_addresses.query(f"address == '{line['address']}'")
                    address_id = query['id'].iloc[0] if len(query.values) != 0 else None

                    if not address_id:
                        logs_accounts.append(
                            {
                                'target': line['account_number'],
                                'success': False,
                                'detail': 'Адрес ПВЗ не найден в базе данных',
                                'value': line['address'],
                                'line': line['line_number'],
                                'orders_type': orders_type,
                                'server': server.name,
                            }
                        )
                        continue

                    data_accounts_to_db.append(
                        {
                            'number': line['account_number'],
                            'name': line['account_name'],
                            'is_active': True,
                            'address_id': address_id,
                            'server_id': server.id,
                        }
                    )

                    logs_accounts.append(
                        {
                            'target': line['account_number'],
                            'success': True,
                            'detail': 'Аккаунт добавлен в базу данных',
                            'value': line['account_number'],
                            'line': line['line_number'],
                            'orders_type': orders_type,
                            'server': server.name,
                        }
                    )

                processed_accounts.append(line['account_number'])

        # Save and recache accounts
        if data_accounts_to_db:
            await Repository.save_records([{'model': OrdersAccountModel, 'records': data_accounts_to_db}])
            db_accounts = await Repository.get_records(OrdersAccountModel)
            df_accounts = pd.DataFrame([{'id': x.id, 'number': x.number} for x in db_accounts])

        # Process orders
        for orders_type, lines in orders.items():

            for line in reversed(lines):

                # Searching order_id (sid) by uuid in files and data verify
                if not line['uuid']:
                    logs_orders.append(
                        {
                            'target': line['uuid'],
                            'sid': None,
                            'success': False,
                            'detail': 'UUID заказа не указан',
                            'value': line['uuid'],
                            'line': line['line_number'],
                            'orders_type': orders_type,
                            'server': server.name,
                        }
                    )
                    continue

                query = df_orders.query(f"wb_uuid == '{line['uuid']}'")
                order_id = query['id'].iloc[0] if len(query.values) != 0 else None

                if order_id:

                    if orders_type == 'collected' and query['dt_collected'].iloc[0]:
                        continue

                else:
                    found_order_in_plan_orders = None

                    for plan_line in plan_orders:

                        if plan_line['uuid'] == line['uuid']:
                            found_order_in_plan_orders = plan_line
                            break

                    if not found_order_in_plan_orders:
                        logs_orders.append(
                            {
                                'target': line['uuid'],
                                'sid': None,
                                'success': False,
                                'detail': 'Заказа нет ни в плане ни в базе. Требуется обработка вручную',
                                'value': line['uuid'],
                                'line': line['line_number'],
                                'orders_type': orders_type,
                                'server': server.name,
                            }
                        )
                        continue

                    if not found_order_in_plan_orders['sid']:
                        logs_orders.append(
                            {
                                'target': found_order_in_plan_orders['sid'],
                                'sid': found_order_in_plan_orders['sid'],
                                'success': False,
                                'detail': 'Не указан SID заказа',
                                'value': found_order_in_plan_orders['sid'],
                                'line': found_order_in_plan_orders['line_number'],
                                'orders_type': 'plan',
                                'server': server.name,
                            }
                        )
                        continue

                    query = df_orders.query(f"id == {found_order_in_plan_orders['sid']}")
                    order_id = query['id'].iloc[0] if len(query.values) != 0 else None

                    if not order_id:
                        logs_orders.append(
                            {
                                'target': found_order_in_plan_orders['sid'],
                                'sid': found_order_in_plan_orders['sid'],
                                'success': False,
                                'detail': 'Заказ не найден по SID',
                                'value': found_order_in_plan_orders['sid'],
                                'line': found_order_in_plan_orders['line_number'],
                                'orders_type': 'plan',
                                'server': server.name,
                            }
                        )
                        continue

                # Refreshing order
                query = df_orders.query(f"id == {order_id}")

                current_is_success = query['is_success'].iloc[0] if len(query.values) != 0 else None
                current_sys_status = query['sys_status'].iloc[0] if len(query.values) != 0 else None

                db_order_id = query['id'].iloc[0] if len(query.values) != 0 else None
                db_wb_status = line['status']
                db_uuid = line['uuid']
                db_description = None
                db_collect_code = line['collect_code']
                db_dt_ordered = detect_date(line['dt_ordered'])
                db_dt_delivered = detect_date(line['dt_delivered'])
                db_dt_collected = detect_date(line['dt_collected'])

                status_model = None
                for db_status in db_statuses:
                    if db_status.full_match and line['status'] == db_status.title:
                        status_model = db_status
                        db_description = db_status.description
                        break
                    elif not db_status.full_match and db_status.title in line['status']:
                        status_model = db_status
                        db_description = db_status.description
                        break

                if not status_model:
                    logs_orders.append(
                        {
                            'target': order_id,
                            'sid': str(db_order_id),
                            'success': False,
                            'detail': 'Неизвестный статус',
                            'value': line['status'],
                            'line': line['line_number'],
                            'orders_type': orders_type,
                            'server': server.name,
                        }
                    )
                    continue

                wb_price = query['wb_price'].iloc[0] if len(query.values) != 0 else None
                if not wb_price:
                    logs_orders.append(
                        {
                            'target': order_id,
                            'sid': str(db_order_id),
                            'success': False,
                            'detail': 'Не указана стоимость заказа в базе данных',
                            'value': wb_price,
                            'line': line['line_number'],
                            'orders_type': orders_type,
                            'server': server.name,
                        }
                    )
                    continue

                try:
                    int(line['price'])
                except:
                    logs_orders.append(
                        {
                            'target': order_id,
                            'sid': str(db_order_id),
                            'success': False,
                            'detail': 'Цена должна быть числом',
                            'value': line['price'],
                            'line': line['line_number'],
                            'orders_type': orders_type,
                            'server': server.name,
                        }
                    )
                    continue

                org_id = query['org_id'].iloc[0] if len(query.values) != 0 else None
                if not org_id:
                    logs_orders.append(
                        {
                            'target': order_id,
                            'sid': str(db_order_id),
                            'success': False,
                            'detail': 'Организация заказа не опознана',
                            'value': org_id,
                            'line': line['line_number'],
                            'orders_type': orders_type,
                            'server': server.name,
                        }
                    )
                    continue

                query = df_accounts.query(f"number == '{line['account_number']}'")
                db_account_id = query['id'].iloc[0] if len(query.values) != 0 else None

                if not db_account_id:
                    logs_orders.append(
                        {
                            'target': line['uuid'],
                            'sid': str(db_order_id),
                            'success': False,
                            'detail': 'Аккаунт заказа не найден в базе данных',
                            'value': line['account_number'],
                            'line': line['line_number'],
                            'orders_type': orders_type,
                            'server': server.name,
                        }
                    )
                    continue

                if current_is_success != status_model.is_success:

                    # Making payment with status
                    if status_model.pay_product:
                        data_payments_to_db.append(
                            {
                                'amount': int(line['price']),
                                'org_id': org_id,
                                'action_id': 3,
                                'target_id': 1,
                                'record_id': db_order_id,
                            }
                        )

                        logs_payments.append(
                            {
                                'target': db_uuid,
                                'success': True,
                                'detail': 'Оплата стоимости продукта',
                                'value': int(line['price']),
                                'line': line['line_number'],
                                'orders_type': orders_type,
                                'server': server.name,
                            }
                        )

                    if status_model.refund_product:
                        data_payments_to_db.append(
                            {
                                'amount': int(wb_price) if current_sys_status == 2 else int(line['price']),
                                'org_id': org_id,
                                'action_id': 4,
                                'target_id': 1,
                                'record_id': db_order_id,
                            }
                        )

                        logs_payments.append(
                            {
                                'target': db_uuid,
                                'success': True,
                                'detail': 'Разморозка стоимости продукта',
                                'value': int(wb_price) if current_sys_status == 2 else int(line['price']),
                                'line': line['line_number'],
                                'orders_type': orders_type,
                                'server': server.name,
                            }
                        )

                    if status_model.refund_services:

                        service_payments = await Repository.get_records(
                            BalanceHistoryModel,
                            filters=[
                                BalanceHistoryModel.target_id.in_([2, 3, 4]),
                                BalanceHistoryModel.record_id == db_order_id,
                                BalanceHistoryModel.action_id.in_([2, 3])
                            ]
                        )

                        for service_payment in service_payments:
                            data_payments_to_db.append(
                                {
                                    'amount': service_payment.amount,
                                    'org_id': service_payment.org_id,
                                    'action_id': 4 if service_payment.action_id == 2 else 1,
                                    'target_id': service_payment.target_id,
                                    'record_id': service_payment.record_id,
                                }
                            )

                            logs_payments.append(
                                {
                                    'target': db_uuid,
                                    'success': True,
                                    'detail': f'Возврат стоимости услуги (возврат платежа #{service_payment.id})',
                                    'value': service_payment.amount,
                                    'line': line['line_number'],
                                    'orders_type': orders_type,
                                    'server': server.name,
                                }
                            )

                # data insert/update
                data_orders_to_db.append(
                    {
                        'id': db_order_id,
                        'status': status_model.status_number,
                        'wb_status': db_wb_status,
                        'wb_uuid': db_uuid,
                        'description': db_description,
                        'wb_collect_code': db_collect_code,
                        'dt_ordered': db_dt_ordered,
                        'dt_delivered': db_dt_delivered,
                        'dt_collected': db_dt_collected,
                        'account_id': db_account_id,
                        'picker_status_id': status_model.id
                    }
                )

                logs_orders.append(
                    {
                        'target': db_uuid,
                        'sid': str(db_order_id),
                        'success': True,
                        'detail': 'Данные заказа обновлены в базе данных',
                        'value': db_uuid,
                        'line': line['line_number'],
                        'orders_type': orders_type,
                        'server': server.name,
                    }
                )

                processed_orders.append(str(db_order_id))

        # Checking plan to refund bad orders
        for line in plan_orders:

            if not line['sid']:
                logs_orders.append(
                    {
                        'target': line['sid'],
                        'sid': line['sid'],
                        'success': False,
                        'detail': 'Не указан SID заказа',
                        'value': line['sid'],
                        'line': line['line_number'],
                        'orders_type': 'plan',
                        'server': server.name,
                    }
                )
                continue

            if not line['status']:
                logs_orders.append(
                    {
                        'target': line['sid'],
                        'sid': line['sid'],
                        'success': False,
                        'detail': 'Не указан статус заказа',
                        'value': line['status'],
                        'line': line['line_number'],
                        'orders_type': 'plan',
                        'server': server.name,
                    }
                )
                continue

            if 'Все артикулы заказаны' not in line['status'] and line['sid'] not in processed_orders:

                query = df_orders.query(f"id == {line['sid']}")
                order_id = query['id'].iloc[0] if len(query.values) != 0 else None

                if not order_id:
                    logs_orders.append(
                        {
                            'target': line['sid'],
                            'sid': line['sid'],
                            'success': False,
                            'detail': 'Заказ не найден по SID',
                            'value': line['sid'],
                            'line': line['line_number'],
                            'orders_type': 'plan',
                            'server': server.name,
                        }
                    )
                    continue

                wb_price = query['wb_price'].iloc[0] if len(query.values) != 0 else None
                if not wb_price:
                    logs_orders.append(
                        {
                            'target': order_id,
                            'sid': line['sid'],
                            'success': False,
                            'detail': 'Не указана стоимость заказа в базе данных',
                            'value': wb_price,
                            'line': line['line_number'],
                            'orders_type': 'plan',
                            'server': server.name,
                        }
                    )
                    continue

                org_id = query['org_id'].iloc[0] if len(query.values) != 0 else None
                if not org_id:
                    logs_orders.append(
                        {
                            'target': order_id,
                            'sid': line['sid'],
                            'success': False,
                            'detail': 'Организация заказа не опознана',
                            'value': org_id,
                            'line': line['line_number'],
                            'orders_type': 'plan',
                            'server': server.name,
                        }
                    )
                    continue

                data_payments_to_db.append(
                    {
                        'amount': int(wb_price),
                        'org_id': org_id,
                        'action_id': 4,
                        'target_id': 1,
                        'record_id': order_id,
                    }
                )

                logs_payments.append(
                    {
                        'target': line['sid'],
                        'success': True,
                        'detail': 'Разморозка стоимости продукта',
                        'value': int(wb_price),
                        'line': line['line_number'],
                        'orders_type': 'plan',
                        'server': server.name,
                    }
                )

                service_payments = await Repository.get_records(
                    BalanceHistoryModel,
                    filters=[
                        BalanceHistoryModel.target_id.in_([2, 3, 4]),
                        BalanceHistoryModel.record_id == order_id,
                        BalanceHistoryModel.action_id.in_([2, 3])
                    ]
                )

                for service_payment in service_payments:
                    data_payments_to_db.append(
                        {
                            'amount': service_payment.amount,
                            'org_id': service_payment.org_id,
                            'action_id': 4 if service_payment.action_id == 2 else 1,
                            'target_id': service_payment.target_id,
                            'record_id': service_payment.record_id,
                        }
                    )

                    logs_payments.append(
                        {
                            'target': line['sid'],
                            'success': True,
                            'detail': f'Возврат стоимости услуги (возврат платежа #{service_payment.id})',
                            'value': service_payment.amount,
                            'line': line['line_number'],
                            'orders_type': 'plan',
                            'server': server.name,
                        }
                    )

                data_orders_to_db.append(
                    {
                        'id': order_id,
                        'status': 6,
                        'wb_status': line['status'],
                        'description': line['status'],
                    }
                )

                logs_orders.append(
                    {
                        'target': line['sid'],
                        'sid': line['sid'],
                        'success': True,
                        'detail': 'Данные заказа обновлены в базе данных (Отмена)',
                        'value': line['uuid'],
                        'line': line['line_number'],
                        'orders_type': 'plan',
                        'server': server.name,
                    }
                )

    for item in logs_accounts:
        if not item['success']:
            return {'accounts': logs_accounts, 'orders': logs_orders, 'payments': logs_payments}

    for item in logs_orders:
        if not item['success']:
            return {'accounts': logs_accounts, 'orders': logs_orders, 'payments': logs_payments}

    records_to_db = [
        {'model': OrdersOrderModel, 'records': data_orders_to_db},
        {'model': BalanceHistoryModel, 'records': data_payments_to_db},
    ]

    await Repository.save_records(records_to_db)

    return {'accounts': logs_accounts, 'orders': logs_orders, 'payments': logs_payments}


async def generate_plan_xlsx_2(servers, date):
    for server in servers:
        db_tasks = await Repository.get_records(
            model=OrdersOrderModel,
            filters=[OrdersOrderModel.status == 2, OrdersOrderModel.dt_planed == date],
            filtration=[OrganizationModel.server_id == server.id],
            select_related=[OrdersOrderModel.size],
            deep_related=[
                [OrdersOrderModel.size, ProductSizeModel.product],
                [OrdersOrderModel.size, ProductSizeModel.product, ProductModel.organization]
            ],
            joins=[ProductSizeModel, ProductModel, OrganizationModel],
        )

        result_book = openpyxl.Workbook()

        ws_res = result_book.active
        ws_res['B1'] = 'артикул'
        ws_res['C1'] = 'кол-во'
        ws_res['D1'] = 'размер'
        ws_res['E1'] = 'ключевой запрос'
        ws_res['F1'] = 'вход'
        ws_res['G1'] = 'ПВЗ'
        ws_res['H1'] = 'подсказка'
        ws_res['I1'] = 'порядок'
        ws_res['K1'] = 'статус'
        ws_res['L1'] = 'аккаунт'
        ws_res['M1'] = 'id заказа'
        ws_res['N1'] = 'ип'
        ws_res['O1'] = 'цена'
        ws_res['P1'] = 'sid'
        ws_res['Z2'] = date.strftime('%d.%m')

        ln = 1
        for task in db_tasks:
            ln += 1
            ws_res[f'A{ln}'] = ln - 1
            ws_res[f'B{ln}'] = task.size.product.wb_article
            ws_res[f'C{ln}'] = 1
            ws_res[f'D{ln}'] = task.size.wb_size_origName
            ws_res[f'E{ln}'] = task.wb_keyword
            ws_res[f'N{ln}'] = task.size.product.organization.title
            ws_res[f'O{ln}'] = task.size.wb_price / 100
            ws_res[f'P{ln}'] = task.id

        result_bytes = io.BytesIO()
        result_book.save(result_bytes)
        result_bytes.seek(0)

        fn = f'{Strings.alphanumeric(32)}.xlsx'
        await Repository.s3_save(result_bytes.getvalue(), fn)

        await Repository.save_records([
            {
                'model': PickerHistoryModel, 'records':
                [
                    {
                        'result': fn,
                        'logs': '-',
                        'server_id': server.id,
                    }
                ]
            }
        ])


async def generate_plan_xlsx(servers, bad_accounts):
    class ScheduleCell:
        dt: datetime.datetime
        org: str | None

        def __init__(self, dt) -> None:
            self.dt = dt
            self.org = None

    def calc_cells(dts, dte, osd):
        return int(
            (dte - dts).total_seconds() /
            osd.total_seconds()
        )

    async def proc_0(org_accs, used_addrs):
        new_org_accs = []

        for i in range(0, len(org_accs)):

            # K - amount of orders planned on account's address for today
            K = used_addrs.count(org_accs[i]['address_id'])
            org_accs[i]['K'] = K

            # L = K + W
            org_accs[i]['L'] = K + org_accs[i]['W']

            # Check L
            if org_accs[i]['L'] in range(org_accs[i]['contractor'].load_l_min,
                                         org_accs[i]['contractor'].load_l_max + 1):
                new_org_accs.append(org_accs[i])

        return new_org_accs

    async def proc_1(accs, settings):
        maxH = max(acc['H'] for acc in accs)
        maxH = 100 if maxH == 0 else maxH

        for acc in accs:

            acc['X'] = ((maxH - acc['H']) / maxH) * 100
            acc['Z'] = 0
            acc['AB'] = acc['L']
            acc['AD'] = settings.r4 - acc['T']

            if acc['M'] is not None:

                if acc['H'] == 0:
                    acc['Z'] = settings.r2

                else:

                    t = (datetime.datetime.now() - acc['M']).days

                    if t > settings.r2:
                        acc['Z'] = settings.r2 + t * settings.r3

                    else:
                        acc['Z'] = t

            else:
                acc['Z'] = settings.r2 + 10

            if acc['L'] == 0:
                acc['AB'] = 0

            elif acc['L'] >= settings.r4:
                acc['AB'] = settings.r4 - 1

            if acc['T'] == 0:
                acc['AD'] = 0

        return accs

    async def proc_2(accs, settings):
        maxX = max(acc['X'] for acc in accs)
        maxZ = max(acc['Z'] for acc in accs)

        for acc in accs:
            if not acc['M']:
                maxZ = settings.r2 + 10
                break

        maxX = 100 if maxX == 0 else maxX
        maxZ = 100 if maxZ == 0 else maxZ

        for acc in accs:
            acc['Y'] = (acc['X'] / maxX) * 100
            acc['AA'] = (acc['Z'] / maxZ) * 100
            acc['AC'] = (acc['AB'] / (settings.r4 - 1)) * 100
            acc['AE'] = acc['AD'] / settings.r4 * 100

        return accs

    async def proc_3(accs, settings):
        maxY = max(acc['Y'] for acc in accs)
        maxAA = max(acc['AA'] for acc in accs)
        maxAC = max(acc['AC'] for acc in accs)
        maxAE = max(acc['AE'] for acc in accs)

        for acc in accs:
            acc['AF'] = (
                    (
                            acc['Y'] * settings.l2 +
                            acc['AA'] * settings.l3 +
                            acc['AC'] * settings.l4 +
                            acc['AE'] * settings.l5
                    ) * 100 /
                    (
                            maxY * settings.l2 +
                            maxAA * settings.l3 +
                            maxAC * settings.l4 +
                            maxAE * settings.l5
                    )
            )

        return accs

    bad_accounts = (
        bad_accounts.replace('\n', '').replace(' ', '').split('\r')
        if bad_accounts
        else []
    )

    db_settings = await DefaultRepository.get_records(PickerSettingsModel)
    db_settings = db_settings[0]

    db_orders = await DefaultRepository.get_records(
        OrdersOrderModel,
        filters=[OrdersOrderModel.dt_ordered.isnot(None)],
        select_related=[OrdersOrderModel.size, OrdersOrderModel.account],
        deep_related=[
            [OrdersOrderModel.size, ProductSizeModel.product],
            [OrdersOrderModel.size, ProductSizeModel.product, ProductModel.organization]
        ],
        joins=[ProductSizeModel, ProductModel, OrganizationModel],
    )

    for order in db_orders:
        if not order.account:
            raise Exception(f'Не указан аккаунт у заказа #{order.id}')

    df_orders = pd.DataFrame(
        {
            'account_id': order.account_id,
            'address_id': order.account.address_id,
            'org_id': order.size.product.org_id,
            'org': order.size.product.organization.title,
            'dt_ordered': order.dt_ordered,
            'dt_delivered': order.dt_delivered,
            'dt_collected': order.dt_collected,
        }
        for order in db_orders
    )

    settings_last_order = datetime.timedelta(days=db_settings.lo)
    settings_account_life = datetime.timedelta(days=db_settings.al)
    settings_now = datetime.datetime.now()
    settings_now_date = datetime.date.today()
    settings_k_format = db_settings.k_format

    used_addresses = []
    total_selected_accounts = []

    for server in servers:

        db_tasks = await DefaultRepository.get_records(
            model=OrdersOrderModel,
            filters=[OrdersOrderModel.status == 2],
            filtration=[OrganizationModel.server_id == server.id],
            select_related=[OrdersOrderModel.size],
            deep_related=[
                [OrdersOrderModel.size, ProductSizeModel.product],
                [OrdersOrderModel.size, ProductSizeModel.product, ProductModel.organization]
            ],
            joins=[ProductSizeModel, ProductModel, OrganizationModel],
        )

        selected_accs, used_addrs, accs = [], used_addresses.copy(), []
        used_accs = {c.contractor.name: [] for c in server.contractors}

        logs_book = openpyxl.Workbook()
        result_book = openpyxl.Workbook()

        ws_res = result_book.active
        ws_res['B1'] = 'артикул'
        ws_res['C1'] = 'кол-во'
        ws_res['D1'] = 'размер'
        ws_res['E1'] = 'ключевой запрос'
        ws_res['F1'] = 'вход'
        ws_res['G1'] = 'ПВЗ'
        ws_res['H1'] = 'подсказка'
        ws_res['I1'] = 'порядок'
        ws_res['K1'] = 'статус'
        ws_res['L1'] = 'аккаунт'
        ws_res['M1'] = 'id заказа'
        ws_res['N1'] = 'ип'
        ws_res['O1'] = 'цена'
        ws_res['P1'] = 'sid'
        ws_res['Z2'] = settings_now.strftime('%d.%m')

        line = 1
        for order in db_tasks:
            line += 1

            size_name = ''
            if order.size.wb_size_name:
                size_name = order.size.wb_size_name
            if order.size.wb_size_origName:
                size_name = order.size.wb_size_origName

            ws_res[f'A{line}'] = line - 1
            ws_res[f'B{line}'] = order.size.product.wb_article
            ws_res[f'C{line}'] = 1
            ws_res[f'D{line}'] = size_name
            ws_res[f'E{line}'] = order.wb_keyword
            ws_res[f'N{line}'] = order.size.product.organization.title
            ws_res[f'O{line}'] = order.size.wb_price // 100
            ws_res[f'P{line}'] = order.id

        names = [(order.size.product.organization.title, order.size.product.wb_article) for order in db_tasks]

        arts = {}
        orgs = {}

        for org, art in names:
            if orgs.get(org):
                orgs[org] += 1

            else:
                orgs[org] = 1
                arts[org] = {}

            if arts[org].get(art):
                arts[org][art] += 1

            else:
                arts[org][art] = 1

        orgs = sorted(orgs.items(), key=operator.itemgetter(1), reverse=True)
        orgs = dict(orgs)

        ################################################################################################################
        # SCHEDULE
        ################################################################################################################

        logs = logs_book.create_sheet('Расписание')
        logs['A1'] = 'Количество задач'
        logs['A2'] = 'Количество ячеек'
        logs['A3'] = 'Шаг ячеек (мин)'
        logs['A4'] = 'Время начала'
        logs['A5'] = 'Время окончания'
        logs['A6'] = 'Размер созданного расписния'
        logs['A7'] = 'Время первой ячейки'
        logs['A8'] = 'Время последней ячейки'
        logs['A9'] = 'Название расписания'
        logs['B9'] = server.schedule.title

        tasks_amount = sum(orgs.values())
        max_one_step_duration = server.schedule.time_max_min_per_step
        min_one_step_duration = server.schedule.time_min_min_per_step

        start_time = datetime.datetime.combine(settings_now_date, server.schedule.time_start)
        end_time = datetime.datetime.combine(settings_now_date, server.schedule.time_end)

        middle_point = datetime.datetime.combine(settings_now_date, server.schedule.time_first_point)
        last_point = datetime.datetime.combine(settings_now_date, server.schedule.time_second_point)

        # end time point
        one_step_duration = datetime.timedelta(minutes=max_one_step_duration)
        cells_amount = calc_cells(start_time, end_time, one_step_duration)

        # middle point
        if cells_amount < tasks_amount:
            end_time = middle_point
            cells_amount = calc_cells(start_time, end_time, one_step_duration)

            # middle point + dynamic
            if cells_amount < tasks_amount:

                one_step_duration = datetime.timedelta(minutes=(cells_amount / tasks_amount) * max_one_step_duration)
                one_step_duration = one_step_duration - datetime.timedelta(microseconds=one_step_duration.microseconds)

                # last point
                if one_step_duration < datetime.timedelta(minutes=min_one_step_duration):

                    one_step_duration = datetime.timedelta(minutes=max_one_step_duration)
                    end_time = last_point
                    cells_amount = calc_cells(start_time, end_time, one_step_duration)

                    # last point + dynamic
                    if cells_amount < tasks_amount:
                        one_step_duration = datetime.timedelta(
                            minutes=(cells_amount / tasks_amount) * max_one_step_duration)
                        one_step_duration = one_step_duration - datetime.timedelta(
                            microseconds=one_step_duration.microseconds)
                        cells_amount = calc_cells(start_time, end_time, one_step_duration)

                        if one_step_duration < datetime.timedelta(minutes=min_one_step_duration):
                            logs['A10'] = f'Недостаточно ячеек расписания для {server.name}'

                            logs_bytes = io.BytesIO()

                            logs_book.save(logs_bytes)

                            logs_bytes.seek(0)

                            logs_filename, logs_filetype = await s3_save(logs_bytes.getvalue(), generate_filename(),
                                                                         'xlsx')

                            await DefaultRepository.save_records([
                                {
                                    'model': PickerHistoryModel, 'records':
                                    [
                                        {
                                            'logs': f'{logs_filename}.{logs_filetype}',
                                            'server_id': server.id,
                                        }
                                    ]
                                }
                            ])

                            continue

        logs['B1'] = tasks_amount
        logs['B2'] = cells_amount
        logs['B3'] = one_step_duration
        logs['B4'] = start_time
        logs['B5'] = end_time

        # Creating schedule
        schedule = []
        current_time = start_time
        while current_time < end_time:
            schedule.append(ScheduleCell(current_time))
            current_time += one_step_duration

        if schedule[-1].dt > end_time - one_step_duration:
            schedule = schedule[:-1]

        logs['B6'] = f'{len(schedule)} ячеек'
        logs['B7'] = f'{schedule[0].dt}'
        logs['B8'] = f'{schedule[-1].dt}'

        time_length = (end_time - start_time).total_seconds()

        # Assigning organizations to schedule
        for org in orgs:

            amount = orgs[org]
            single_length = time_length / amount
            assigned_amount = 0

            for x in range(amount):
                dt = start_time + datetime.timedelta(seconds=x * single_length)

                # Find closest to dt free cell in schedule
                closest_cell = None

                for cell in schedule:
                    if cell.dt >= dt and not cell.org:
                        closest_cell = cell
                        break

                if closest_cell:
                    assigned_amount += 1
                    closest_cell.org = org

        used_dt = []
        used_lines = []

        for org in arts:
            m = max(arts[org].values())

            for i in range(m):

                for art in arts[org]:

                    if arts[org][art] != 0:

                        line = 1
                        for row in ws_res.rows:

                            if str(row[1].value) == str(art) and str(row[13].value) == str(
                                    org) and line not in used_lines:

                                for cell in schedule:
                                    if cell.org and cell.org == org and cell.dt not in used_dt:
                                        used_dt.append(cell.dt)
                                        row[10].value = cell.dt.strftime(settings_k_format)
                                        break

                                used_lines.append(line)
                                break

                            line += 1

                        arts[org][art] -= 1

        ################################################################################################################
        # COMMON POOL
        ################################################################################################################
        db_accounts = await DefaultRepository.get_records(
            OrdersAccountModel,
            filters=[OrdersAccountModel.server_id == server.id],
            select_related=[OrdersAccountModel.address],
        )

        common_pool = []
        logs = logs_book.create_sheet('Общий пул')
        line = 1
        logs[f'A{line}'] = 'ID аккаунта'
        logs[f'B{line}'] = 'Номер'
        logs[f'C{line}'] = 'Имя'
        logs[f'D{line}'] = 'Статус аккаунта'
        logs[f'E{line}'] = 'ID адреса'
        logs[f'F{line}'] = 'Адрес'
        logs[f'G{line}'] = 'Район'
        logs[f'H{line}'] = 'Курьер'
        logs[f'I{line}'] = 'Статус адреса'
        logs[f'J{line}'] = 'Исключен'
        logs[f'K{line}'] = 'T'
        logs[f'L{line}'] = 'W'

        for account in db_accounts:

            line += 1
            logs[f'A{line}'] = account.id
            logs[f'B{line}'] = account.number
            logs[f'C{line}'] = account.name
            logs[f'D{line}'] = account.number in bad_accounts or account.is_active
            logs[f'E{line}'] = account.address.id
            logs[f'F{line}'] = account.address.address
            logs[f'G{line}'] = account.address.district
            logs[f'H{line}'] = account.address.contractor
            logs[f'I{line}'] = account.address.is_active

            # account active, account address active, account number not in bad_accounts
            if account.number in bad_accounts:
                logs[f'J{line}'] = 'Исключен, реестровый аккаунт'
                continue

            if not account.is_active:
                logs[f'J{line}'] = 'Исключен, аккаунт не активен'
                continue

            if not account.address.is_active:
                logs[f'J{line}'] = 'Исключен, адрес аккаунта не активен'
                continue

            # account last order
            query = (
                df_orders
                .query(f"account_id == {account.id} and dt_collected.isnull()")
                .sort_values('dt_ordered', ascending=True)
                .head(1)
            )

            last_order = query['dt_ordered'].iloc[0] if len(query.values) != 0 else None

            if last_order and last_order + settings_last_order < settings_now_date:
                logs[f'J{line}'] = \
                    f'Исключен, последний заказ на аккаунте позднее {settings_last_order.days} дней с текущей даты'
                continue

            # account registration date
            query = (
                df_orders
                .query(f"account_id == {account.id}")
                .sort_values('dt_ordered', ascending=True)
                .head(1)
            )

            default_reg_date = query['dt_ordered'].iloc[0] if len(query.values) != 0 else settings_now_date
            reg_date = account.reg_date if account.reg_date else default_reg_date

            if reg_date + settings_account_life < settings_now_date:
                logs[
                    f'J{line}'] = f'Исключен, с даты регистрации аккаунта {reg_date} прошло ' \
                                  f'{(settings_now_date - reg_date).days} ' \
                                  f'дней с текущей даты (макс допустимо {settings_account_life.days} дней)'
                continue

            # Contractor
            contractor = next(
                (
                    server_contractor for server_contractor in server.contractors
                    if server_contractor.contractor.id == account.address.contractor_id
                ), None
            )

            if not contractor:
                logs[f'J{line}'] = 'Исключен, курьер адреса аккаунта отсутствует на сервере'
                continue

            # T
            T = (df_orders.query(f'account_id == {account.id} and dt_collected.isnull()').shape[0])

            if T not in range(contractor.load_t_min, contractor.load_t_max + 1):
                logs[f'J{line}'] = 'Исключен, выход за допустимый интервал кол-ва активных заказов на аккаунте'
                continue

            # W
            W = (df_orders.query(f'address_id == {account.address_id} and dt_collected.isnull()').shape[0])

            common_pool.append({
                'account_id': account.id,
                'address_id': account.address_id,
                'number': account.number,
                'address': account.address.address,
                'contractor': contractor,
                'district': account.address.district,

                'T': T,  # amount of active orders on account
                'W': W,  # amount of active orders on account's address
            })

            logs[f'J{line}'] = 'Не исключен'
            logs[f'K{line}'] = T
            logs[f'L{line}'] = W

        ################################################################################################################
        # ACCOUNTS PICKER
        ################################################################################################################

        for org in orgs:
            amount = orgs[org]

            org_used_accs, org_used_addr, org_accs = [], [], []

            cs_tmp = sorted(server.contractors, key=lambda x: x.load_percent, reverse=True)
            contractors = []
            remaining = amount

            for c_tmp in cs_tmp[:-1]:

                cmax = round(amount * c_tmp.load_percent)

                if cmax >= remaining:
                    cmax = remaining
                    remaining = 0

                elif cmax < remaining:
                    remaining -= cmax

                contractors.append({
                    'name': c_tmp.contractor.name,
                    'usages': [],
                    'max': cmax
                })

            contractors.append({
                'name': cs_tmp[-1].contractor.name,
                'usages': [],
                'max': remaining
            })

            ws_org = logs_book.create_sheet(org)
            line = 1
            ws_org[f'A{line}'] = 'ID аккаунта'
            ws_org[f'B{line}'] = 'Номер'
            ws_org[f'C{line}'] = 'ID адреса'
            ws_org[f'D{line}'] = 'Адрес'
            ws_org[f'E{line}'] = 'Курьер'
            ws_org[f'F{line}'] = 'Район'
            ws_org[f'G{line}'] = 'Организация'
            ws_org[f'H{line}'] = 'Исключен'
            ws_org[f'I{line}'] = 'T'
            ws_org[f'J{line}'] = 'W'
            ws_org[f'K{line}'] = 'I'
            ws_org[f'L{line}'] = 'M'
            ws_org[f'M{line}'] = 'H'
            ws_org[f'N{line}'] = 'K'
            ws_org[f'O{line}'] = 'L'
            ws_org[f'P{line}'] = 'X'
            ws_org[f'Q{line}'] = 'Z'
            ws_org[f'R{line}'] = 'AB'
            ws_org[f'S{line}'] = 'AD'
            ws_org[f'T{line}'] = 'Y'
            ws_org[f'U{line}'] = 'AA'
            ws_org[f'V{line}'] = 'AC'
            ws_org[f'W{line}'] = 'AE'
            ws_org[f'X{line}'] = 'AF (Итоговый)'
            ws_org[f'Y{line}'] = 'Выбран'

            for acc in common_pool:

                line += 1

                ws_org[f'A{line}'] = acc['account_id']
                ws_org[f'B{line}'] = acc['number']
                ws_org[f'C{line}'] = acc['address_id']
                ws_org[f'D{line}'] = acc['address']
                ws_org[f'E{line}'] = acc['contractor'].contractor.name
                ws_org[f'F{line}'] = acc['district']
                ws_org[f'G{line}'] = org
                ws_org[f'I{line}'] = acc['T']
                ws_org[f'J{line}'] = acc['W']

                I = (df_orders.query(f"account_id == {acc['account_id']} and org == '{org}'").shape[0])
                ws_org[f'K{line}'] = I

                if I != 0:
                    ws_org[f'H{line}'] = 'Исключен, ИП уже заказывал на этот аккаунт'
                    continue

                query = (
                    df_orders
                    .query(f"address_id == {acc['address_id']} and org == '{org}'")
                    .sort_values('dt_ordered', ascending=False)
                    .head(1)
                )

                M = query['dt_ordered'].iloc[0] if len(query.values) != 0 else None

                if M:
                    ws_org[f'L{line}'] = M.strftime('%d.%m.%Y')
                else:
                    ws_org[f'L{line}'] = 'Не найдено'

                if M and M > acc['contractor'].load_m:
                    ws_org[f'H{line}'] = 'Исключен, дата последнего заказа вне допустимого интервала'
                    continue

                H = (df_orders.query(f"address_id == {acc['address_id']} and org == '{org}'").shape[0])

                ws_org[f'M{line}'] = H
                org_accs.append({
                    'org': org,
                    **acc,
                    'I': I,
                    'H': H,  # amount of all orders of org on address of account
                    'M': M,  # date of the last purchase of org on account's address
                    'logs': line,
                })

                ws_org[f'H{line}'] = 'Не исключен'

            for i in range(0, amount):

                org_accs = await proc_0(org_accs, used_addrs)  # L | K W

                if not org_accs:
                    break

                org_accs = await proc_1(org_accs, db_settings)  # X Z AB AD   | H L T
                org_accs = await proc_2(org_accs, db_settings)  # Y AA AC AE  | X Z AB AD
                org_accs = await proc_3(org_accs, db_settings)  # AF          | Y AA AC AE

                org_accs = sorted(org_accs, key=lambda x: x['AF'], reverse=True)

                if i == 0:
                    for oa in org_accs:
                        ws_org[f'N{oa["logs"]}'] = oa['K']
                        ws_org[f'O{oa["logs"]}'] = oa['L']
                        ws_org[f'P{oa["logs"]}'] = oa['X']
                        ws_org[f'Q{oa["logs"]}'] = oa['Z']
                        ws_org[f'R{oa["logs"]}'] = oa['AB']
                        ws_org[f'S{oa["logs"]}'] = oa['AD']
                        ws_org[f'T{oa["logs"]}'] = oa['Y']
                        ws_org[f'U{oa["logs"]}'] = oa['AA']
                        ws_org[f'V{oa["logs"]}'] = oa['AC']
                        ws_org[f'W{oa["logs"]}'] = oa['AE']
                        ws_org[f'X{oa["logs"]}'] = oa['AF']

                for org_acc in org_accs:

                    if org_acc['address_id'] in org_used_addr:
                        continue

                    # J
                    for contractor in contractors:
                        if (len(contractor['usages']) != contractor['max'] and
                                org_acc['contractor'].contractor.name == contractor['name']):

                            HH = (
                                df_orders.query(f'address_id == {org_acc["address_id"]} and dt_collected.isnull()').loc[
                                :,
                                'account_id'].unique().shape[0])

                            aa_accs = []
                            JJ = 0
                            for aacc in total_selected_accounts:
                                if aacc['address_id'] == org_acc['address_id'] and aacc['account_id'] not in aa_accs:
                                    aa_accs.append(aacc['account_id'])
                                    JJ += 1

                            II = HH + JJ  # кол во аккаунтов с активными заказами на адресе аккаунта

                            if II >= org_acc['contractor'].load_i and org_acc['T'] == 0:
                                continue

                            J = used_accs[org_acc['contractor'].contractor.name].count(org_acc['account_id'])

                            if J in range(org_acc['contractor'].load_j_min, org_acc['contractor'].load_j_max + 1):
                                org_acc['J'] = J
                                contractor['usages'].append(org_acc['account_id'])

                                used_addrs.append(org_acc['address_id'])
                                used_accs[org_acc['contractor'].contractor.name].append(org_acc['account_id'])

                                org_used_addr.append(org_acc['address_id'])
                                selected_accs.append(org_acc)

                                total_selected_accounts.append(org_acc)

                                ws_org[f'Y{org_acc["logs"]}'] = 'Выбран'
                                fill = PatternFill(patternType='solid', fgColor='FF00FF00')
                                ws_org[f'D{org_acc["logs"]}'].fill = fill

                                break

            s = 0
            ws_org[f'AA1'] = f'Всего'
            ws_org[f'AC1'] = f'из {amount}'

            for i in range(0, len(contractors)):
                t = len(contractors[i]['usages'])
                ws_org[f'AA{i + 2}'] = contractors[i]['name']
                ws_org[f'AB{i + 2}'] = t
                ws_org[f'AC{i + 2}'] = contractors[i]['max']
                s += t

            ws_org[f'AB1'] = f'{s}'

        random.shuffle(selected_accs)
        line = 1
        for row in ws_res.rows:
            line += 1
            for acc in selected_accs:

                if acc['org'] == ws_res[f'N{line}'].value and not acc.get('selected'):
                    acc['selected'] = True
                    ws_res[f'G{line}'] = acc['address']
                    ws_res[f'L{line}'] = acc['number']
                    break

        ws_res[f'Z2'] = datetime.datetime.now().strftime('%d.%m')

        result_bytes = io.BytesIO()
        logs_bytes = io.BytesIO()

        result_book.save(result_bytes)
        logs_book.save(logs_bytes)

        result_bytes.seek(0)
        logs_bytes.seek(0)

        result_filename, result_filetype = await s3_save(result_bytes.getvalue(), generate_filename(), 'xlsx')
        logs_filename, logs_filetype = await s3_save(logs_bytes.getvalue(), generate_filename(), 'xlsx')

        await DefaultRepository.save_records([
            {
                'model': PickerHistoryModel, 'records':
                [
                    {
                        'result': f'{result_filename}.{result_filetype}',
                        'logs': f'{logs_filename}.{logs_filetype}',
                        'server_id': server.id,
                    }
                ]
            }
        ])

        used_addresses = used_addrs.copy()